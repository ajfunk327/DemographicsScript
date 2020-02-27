import os
from openpyxl import load_workbook

age = 0;
participants = 0;
fCount = 0;
mCount = 0; 
Caucasian = 0;
AAmerican = 0;
Hispanic = 0;
Other = 0;

participantNum = int(input("Please input the number of participants: "));

rowAge, colAge = input("Please input the row and column for age data, please seperate with a space: ").split();
rowGend, colGend = input("Please input the row and column for gender data, please seperate with a space: ").split();
rowE, colE = input("Please input the row and column for ethnicity data, please seperate with a space: ").split();

print(" ");

for i in range(0, participantNum):
    try:
        wb = load_workbook(filename = 'Participant{}_Demo.xlsx'.format(i))
        sheet = wb.active
        cell = sheet.cell(row = int(rowAge), column = int(colAge)) #2 2
        age = age + cell.value
        participants = participants + 1
        cell = sheet.cell(row = int(rowGend), column = int(colGend)) #1 2
        if cell.value == "Female":
            fCount = fCount + 1
        else:
            mCount = mCount + 1
        cell = sheet.cell(row = int(rowE), column = int(colE)) #3 2
        if cell.value == "Caucasian":
            Caucasian = Caucasian + 1;
        elif cell.value == "A . American":
            AAmerican = AAmerican + 1;
        elif cell.value == "A. American":
            AAmerican = AAmerican + 1;
        elif cell.value == "Hispanic":
            Hispanic = Hispanic + 1;
        else:
            Other = Other + 1;
    except FileNotFoundError:
        print("Participant {} DNE. Moving on...".format(i));

print(" ");

avgAge = age/participants;
print('Number of Participants: {}'.format(participants));
print('Avg age: {}'.format(avgAge));
print(" ");

print('Num of Females: {}. \t Avg: {}'.format(fCount, fCount/participants));
print('Num of Males: {}. \t Avg: {}'.format(mCount, mCount/participants));
print(" ");
print('Check of Female, Male: {}'.format(mCount + fCount));
print(" ");
print('Num of Caucasians: {}'.format(Caucasian));
print('Num of African Americans: {}'.format(AAmerican));
print('Num of Hispanics: {}'.format(Hispanic));
print('Num of Other: {}'.format(Other));
print(" ");
print('Check race num: {}'.format(Caucasian + AAmerican + Hispanic + Other));
